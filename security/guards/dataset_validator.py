# ─────────────────────────────────────────────────────────────────────────────
# guards/dataset_validator.py   (NEW — replaces/extends stub)
#
# Dataset Structure Validator — Layer 4 of the Upload Protection Pipeline
#
# WHAT THIS MODULE DOES:
#   After the file passes extension, MIME, and size checks, we open the file
#   and validate its internal structure:
#
#   CSV:
#     - Can we read it as CSV at all? (not binary garbage)
#     - Does it have headers?
#     - Are rows consistent? (same column count per row)
#     - Is it within column/row limits?
#
#   JSON:
#     - Is it valid JSON syntax?
#     - Is it a list-of-dicts (table) or a single dict?
#     - Is the nesting depth within safe limits?
#     - Does it have too many keys?
#
#   Excel (.xlsx):
#     - Is it a readable workbook?
#     - Does it have at least one sheet?
#     - Is it within row/sheet limits?
#
# WHY STRUCTURE VALIDATION?
#   Corrupted, malformed, or unusually structured files can:
#     - Crash CSV parsers (CSV injection via quoting attacks)
#     - Exhaust memory (deeply nested JSON, huge Excel files)
#     - Cause unexpected behavior in the analytics pipeline
#
# This layer also detects CSV formula injection prefixes (= + - @).
#
# Public API:
#   DatasetValidator
#     validate(file_path, extension) → DatasetValidationReport
#   validate_dataset(file_path, extension) → DatasetValidationReport (convenience)
# ─────────────────────────────────────────────────────────────────────────────

import sys
import csv
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

from configs.upload_settings import (
    CSV_MAX_COLUMNS, CSV_MAX_ROWS,
    JSON_MAX_DEPTH, JSON_MAX_KEYS,
    XLSX_MAX_SHEETS, XLSX_MAX_ROWS,
    CSV_FORMULA_PREFIXES,
    UploadSeverity,
)

logger = logging.getLogger("security.dataset_validator")

# Optional Excel support
try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    logger.warning("openpyxl not installed — XLSX validation unavailable. pip install openpyxl")


# ══════════════════════════════════════════════════════════════════════════════
# RESULT TYPES
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class StructureFinding:
    """A single structural issue found during dataset validation."""
    check      : str    # e.g. "column_count", "json_depth", "formula_injection"
    row        : int    # row index (-1 if not row-specific)
    column     : str    # column name or "" if not column-specific
    detail     : str    # human-readable description
    severity   : str    # UploadSeverity value


@dataclass
class DatasetValidationReport:
    """
    Result of dataset structure validation.

    Fields
    ------
    file_name       : original filename
    format          : "csv" | "json" | "xlsx"
    valid           : True if structure is valid and safe
    row_count       : number of data rows parsed
    column_count    : number of columns (0 for JSON root key count)
    columns         : list of column names (CSV/XLSX)
    findings        : list of StructureFinding objects
    rejection_reason: first finding's detail (or "" if valid)
    severity        : worst severity across findings
    """
    file_name        : str
    format           : str
    valid            : bool
    row_count        : int           = 0
    column_count     : int           = 0
    columns          : List[str]     = field(default_factory=list)
    findings         : List[StructureFinding] = field(default_factory=list)
    rejection_reason : str           = ""
    severity         : str           = UploadSeverity.LOW

    def __bool__(self) -> bool:
        return self.valid

    def to_dict(self) -> dict:
        return {
            "file_name"       : self.file_name,
            "format"          : self.format,
            "valid"           : self.valid,
            "row_count"       : self.row_count,
            "column_count"    : self.column_count,
            "columns"         : self.columns,
            "severity"        : self.severity,
            "rejection_reason": self.rejection_reason,
            "findings_count"  : len(self.findings),
            "findings": [
                {"check": f.check, "row": f.row, "detail": f.detail, "severity": f.severity}
                for f in self.findings
            ],
        }


# ══════════════════════════════════════════════════════════════════════════════
# DATASET VALIDATOR
# ══════════════════════════════════════════════════════════════════════════════

class DatasetValidator:
    """
    Validates the internal structure of uploaded dataset files.

    Checks CSV for readability, consistent columns, row limits, and formula
    injection. Checks JSON for valid syntax, structure, and depth limits.
    Checks XLSX for readable workbook with valid sheets.

    Example
    -------
    >>> dv = DatasetValidator()
    >>> report = dv.validate(Path("data.csv"), ".csv")
    >>> report.valid
    True
    >>> report.row_count
    42
    """

    def validate(self, file_path: Path, extension: str) -> DatasetValidationReport:
        """
        Validate the dataset structure based on file extension.

        Parameters
        ----------
        file_path : Path — path to the uploaded file
        extension : str  — ".csv", ".json", or ".xlsx"

        Returns
        -------
        DatasetValidationReport

        Raises
        ------
        ValueError — if extension is not a supported format
        """
        extension = extension.lower().strip()
        file_path = Path(file_path)

        if extension == ".csv":
            return self._validate_csv(file_path)
        elif extension == ".json":
            return self._validate_json(file_path)
        elif extension == ".xlsx":
            return self._validate_xlsx(file_path)
        else:
            return DatasetValidationReport(
                file_name="", format="unknown", valid=False,
                rejection_reason=f"Unsupported format: '{extension}'",
                severity=UploadSeverity.HIGH,
            )

    # ── CSV Validation ────────────────────────────────────────────────────────

    def _validate_csv(self, file_path: Path) -> DatasetValidationReport:
        """
        Validate CSV structure.

        Checks:
          1. File is readable as CSV (not corrupt/binary)
          2. Has at least one header row
          3. Column count within CSV_MAX_COLUMNS
          4. Each row has consistent column count (catches mismatched rows)
          5. Row count within CSV_MAX_ROWS
          6. No formula injection prefixes in cells (=, +, -, @, tab, CR)
        """
        filename = file_path.name
        findings : List[StructureFinding] = []
        rows_read = 0
        columns   : List[str] = []

        try:
            # Quick binary content check — reject files with null bytes
            # (CSV parsers accept them but they indicate binary/malicious content)
            raw_bytes = file_path.read_bytes()[:4096]
            if b"\x00" in raw_bytes:
                return DatasetValidationReport(
                    file_name=filename, format="csv", valid=False,
                    rejection_reason="CSV file contains null bytes — likely a binary file, not a dataset.",
                    severity=UploadSeverity.HIGH,
                )

            with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)

                # Read header
                try:
                    header = next(reader)
                except StopIteration:
                    return DatasetValidationReport(
                        file_name=filename, format="csv", valid=False,
                        rejection_reason="CSV file is empty — no rows found.",
                        severity=UploadSeverity.MEDIUM,
                    )

                columns = [h.strip() for h in header]
                expected_cols = len(columns)

                if expected_cols == 0:
                    return DatasetValidationReport(
                        file_name=filename, format="csv", valid=False,
                        rejection_reason="CSV header row is empty.",
                        severity=UploadSeverity.MEDIUM,
                    )

                if expected_cols > CSV_MAX_COLUMNS:
                    return DatasetValidationReport(
                        file_name=filename, format="csv", valid=False,
                        column_count=expected_cols, columns=columns,
                        rejection_reason=f"CSV has {expected_cols} columns — exceeds {CSV_MAX_COLUMNS} column limit.",
                        severity=UploadSeverity.HIGH,
                    )

                # Check header column names for formula injection
                for col in columns:
                    finding = self._check_formula_injection(col, row=-1, column="header")
                    if finding:
                        findings.append(finding)

                # Read data rows
                for row_idx, row in enumerate(reader):
                    rows_read += 1

                    if rows_read > CSV_MAX_ROWS:
                        findings.append(StructureFinding(
                            check="row_limit", row=rows_read, column="",
                            detail=f"CSV exceeds {CSV_MAX_ROWS} row limit.",
                            severity=UploadSeverity.HIGH,
                        ))
                        break

                    # Inconsistent column count
                    if len(row) != expected_cols:
                        findings.append(StructureFinding(
                            check="column_consistency", row=row_idx + 1, column="",
                            detail=(
                                f"Row {row_idx + 1} has {len(row)} columns "
                                f"but header has {expected_cols}."
                            ),
                            severity=UploadSeverity.MEDIUM,
                        ))

                    # Formula injection in cells
                    for col_idx, cell in enumerate(row):
                        col_name = columns[col_idx] if col_idx < len(columns) else f"col_{col_idx}"
                        finding = self._check_formula_injection(cell, row=row_idx + 1, column=col_name)
                        if finding:
                            findings.append(finding)

        except UnicodeDecodeError:
            return DatasetValidationReport(
                file_name=filename, format="csv", valid=False,
                rejection_reason="CSV file contains invalid characters — possibly a binary file.",
                severity=UploadSeverity.HIGH,
            )
        except csv.Error as e:
            return DatasetValidationReport(
                file_name=filename, format="csv", valid=False,
                rejection_reason=f"CSV parsing error: {e}",
                severity=UploadSeverity.MEDIUM,
            )
        except Exception as e:
            return DatasetValidationReport(
                file_name=filename, format="csv", valid=False,
                rejection_reason=f"Unexpected error reading CSV: {e}",
                severity=UploadSeverity.HIGH,
            )

        valid   = len(findings) == 0
        worst   = self._worst_severity(findings)
        reason  = findings[0].detail if findings else ""

        logger.info(
            f"CSV validated: file='{filename}' rows={rows_read} "
            f"cols={len(columns)} findings={len(findings)}"
        )
        return DatasetValidationReport(
            file_name=filename, format="csv", valid=valid,
            row_count=rows_read, column_count=len(columns), columns=columns,
            findings=findings, rejection_reason=reason, severity=worst,
        )

    def _check_formula_injection(self, value: str, row: int, column: str) -> Optional[StructureFinding]:
        """
        Check if a CSV cell value starts with a formula injection prefix.

        Excel and Google Sheets execute cells starting with =, +, -, @
        as formulas. If a malicious cell contains =cmd|..., it can execute
        commands when the CSV is opened in a spreadsheet application.

        Reference: OWASP CSV Injection (CWE-1236)
        """
        if value and value[0] in CSV_FORMULA_PREFIXES:
            return StructureFinding(
                check    = "formula_injection",
                row      = row,
                column   = column,
                detail   = (
                    f"Cell at row {row} column '{column}' starts with '{value[0]}' "
                    "which may trigger formula execution in spreadsheet applications."
                ),
                severity = UploadSeverity.HIGH,
            )
        return None

    # ── JSON Validation ───────────────────────────────────────────────────────

    def _validate_json(self, file_path: Path) -> DatasetValidationReport:
        """
        Validate JSON structure.

        Checks:
          1. File is valid JSON (no syntax errors)
          2. Top-level is dict or list (not string/number)
          3. If list — contains dicts (list-of-records format)
          4. Key count within JSON_MAX_KEYS
          5. Nesting depth within JSON_MAX_DEPTH
        """
        filename = file_path.name
        findings : List[StructureFinding] = []

        try:
            content = file_path.read_text(encoding="utf-8", errors="replace")
        except IOError as e:
            return DatasetValidationReport(
                file_name=filename, format="json", valid=False,
                rejection_reason=f"Cannot read JSON file: {e}",
                severity=UploadSeverity.HIGH,
            )

        # Syntax check
        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            return DatasetValidationReport(
                file_name=filename, format="json", valid=False,
                rejection_reason=f"Invalid JSON syntax: {e}",
                severity=UploadSeverity.MEDIUM,
            )

        # Type check
        if not isinstance(data, (dict, list)):
            return DatasetValidationReport(
                file_name=filename, format="json", valid=False,
                rejection_reason=(
                    f"JSON must be a dict or list at the root level. "
                    f"Got: {type(data).__name__}"
                ),
                severity=UploadSeverity.MEDIUM,
            )

        # Depth check
        depth = self._json_depth(data)
        if depth > JSON_MAX_DEPTH:
            findings.append(StructureFinding(
                check="json_depth", row=-1, column="",
                detail=f"JSON nesting depth {depth} exceeds limit of {JSON_MAX_DEPTH}.",
                severity=UploadSeverity.HIGH,
            ))

        # Key count check
        key_count = self._json_key_count(data)
        if key_count > JSON_MAX_KEYS:
            findings.append(StructureFinding(
                check="json_keys", row=-1, column="",
                detail=f"JSON has {key_count} keys — exceeds limit of {JSON_MAX_KEYS}.",
                severity=UploadSeverity.HIGH,
            ))

        # Determine row count and column info
        if isinstance(data, list):
            row_count    = len(data)
            columns_seen = set()
            for item in data:
                if isinstance(item, dict):
                    columns_seen.update(item.keys())
            columns    = sorted(columns_seen)
            col_count  = len(columns)
        else:
            row_count  = 1
            columns    = list(data.keys())
            col_count  = len(columns)

        valid  = len(findings) == 0
        worst  = self._worst_severity(findings)
        reason = findings[0].detail if findings else ""

        logger.info(
            f"JSON validated: file='{filename}' rows={row_count} "
            f"keys={col_count} depth={depth} findings={len(findings)}"
        )
        return DatasetValidationReport(
            file_name=filename, format="json", valid=valid,
            row_count=row_count, column_count=col_count, columns=columns,
            findings=findings, rejection_reason=reason, severity=worst,
        )

    def _json_depth(self, obj, current: int = 1) -> int:
        """Recursively compute JSON nesting depth."""
        if isinstance(obj, dict):
            if not obj:
                return current
            return max(self._json_depth(v, current + 1) for v in obj.values())
        elif isinstance(obj, list):
            if not obj:
                return current
            return max(self._json_depth(item, current + 1) for item in obj)
        return current

    def _json_key_count(self, obj) -> int:
        """Count total keys recursively."""
        if isinstance(obj, dict):
            return len(obj) + sum(self._json_key_count(v) for v in obj.values())
        elif isinstance(obj, list):
            return sum(self._json_key_count(item) for item in obj)
        return 0

    # ── XLSX Validation ───────────────────────────────────────────────────────

    def _validate_xlsx(self, file_path: Path) -> DatasetValidationReport:
        """
        Validate Excel (.xlsx) workbook structure.

        Checks:
          1. openpyxl can open the file (not corrupt or actually a zip)
          2. At least one worksheet exists
          3. Sheet count within XLSX_MAX_SHEETS
          4. Row count within XLSX_MAX_ROWS
        """
        filename = file_path.name

        if not _HAS_OPENPYXL:
            return DatasetValidationReport(
                file_name=filename, format="xlsx", valid=False,
                rejection_reason="openpyxl is not installed — XLSX validation unavailable.",
                severity=UploadSeverity.MEDIUM,
            )

        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as e:
            return DatasetValidationReport(
                file_name=filename, format="xlsx", valid=False,
                rejection_reason=f"Cannot open XLSX workbook: {e}",
                severity=UploadSeverity.MEDIUM,
            )

        findings : List[StructureFinding] = []

        if not wb.sheetnames:
            return DatasetValidationReport(
                file_name=filename, format="xlsx", valid=False,
                rejection_reason="XLSX workbook has no sheets.",
                severity=UploadSeverity.MEDIUM,
            )

        if len(wb.sheetnames) > XLSX_MAX_SHEETS:
            findings.append(StructureFinding(
                check="sheet_count", row=-1, column="",
                detail=f"Workbook has {len(wb.sheetnames)} sheets — exceeds {XLSX_MAX_SHEETS} sheet limit.",
                severity=UploadSeverity.HIGH,
            ))

        # Inspect first sheet
        ws        = wb.active
        row_count = 0
        columns   : List[str] = []

        try:
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    columns = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                row_count += 1
                if row_count > XLSX_MAX_ROWS:
                    findings.append(StructureFinding(
                        check="row_limit", row=row_count, column="",
                        detail=f"XLSX exceeds {XLSX_MAX_ROWS} row limit.",
                        severity=UploadSeverity.HIGH,
                    ))
                    break
        except Exception as e:
            findings.append(StructureFinding(
                check="read_error", row=-1, column="",
                detail=f"Error reading XLSX rows: {e}",
                severity=UploadSeverity.MEDIUM,
            ))

        wb.close()

        valid  = len(findings) == 0
        worst  = self._worst_severity(findings)
        reason = findings[0].detail if findings else ""

        logger.info(
            f"XLSX validated: file='{filename}' sheets={len(wb.sheetnames)} "
            f"rows={row_count} cols={len(columns)} findings={len(findings)}"
        )
        return DatasetValidationReport(
            file_name=filename, format="xlsx", valid=valid,
            row_count=row_count, column_count=len(columns), columns=columns,
            findings=findings, rejection_reason=reason, severity=worst,
        )

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _worst_severity(self, findings: List[StructureFinding]) -> str:
        order = {
            UploadSeverity.CRITICAL: 4,
            UploadSeverity.HIGH    : 3,
            UploadSeverity.MEDIUM  : 2,
            UploadSeverity.LOW     : 1,
        }
        if not findings:
            return UploadSeverity.LOW
        return max(findings, key=lambda f: order.get(f.severity, 0)).severity


# ── Module-level singleton + convenience ──────────────────────────────────────

_validator_instance: Optional[DatasetValidator] = None


def get_dataset_validator() -> DatasetValidator:
    """Return the module-level singleton DatasetValidator."""
    global _validator_instance
    if _validator_instance is None:
        _validator_instance = DatasetValidator()
    return _validator_instance


def validate_dataset(file_path: Path, extension: str) -> DatasetValidationReport:
    """Module-level convenience: validate dataset structure."""
    return get_dataset_validator().validate(file_path, extension)
